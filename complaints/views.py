import secrets
from datetime import timedelta
from functools import wraps

from django.contrib import messages
from django.contrib.auth import get_user_model, views as auth_views
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Q
from django.db.models.functions import TruncDate
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone

from .forms import (
    BranchAdminForm,
    CityAdminForm,
    ComplaintDetailItemForm,
    ComplaintSourceForm,
    ComplaintSubmissionForm,
    ComplaintUpdateForm,
    SatisfactionRatingForm,
    SiteSettingsForm,
    StaffAccountCreateForm,
    StaffAccountEditForm,
    StaffPhoneEditForm,
    StatusCheckForm,
    StyledPasswordChangeForm,
)
from .models import (
    Branch, City, Complaint, ComplaintDetailItem, ComplaintSource,
    SiteSettings, StaffProfile,
)

User = get_user_model()


def admin_pusat_required(view_func):
    """Membatasi akses hanya untuk user dengan peran Admin Pusat."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        profile = getattr(request.user, 'staff_profile', None)
        if profile is None or not profile.is_admin_pusat:
            raise PermissionDenied('Hanya Admin Pusat yang dapat mengakses halaman ini.')
        return view_func(request, *args, **kwargs)
    return wrapper


def input_staff_required(view_func):
    """Membatasi akses HANYA untuk user dengan peran Staff Input Komplain."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        profile = getattr(request.user, 'staff_profile', None)
        if profile is None or not profile.is_input_staff:
            raise PermissionDenied('Hanya Staff Input Komplain yang dapat mengakses halaman ini.')
        return view_func(request, *args, **kwargs)
    return wrapper


def manager_required(view_func):
    """Membatasi akses HANYA untuk user dengan peran Manager Area."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        profile = getattr(request.user, 'staff_profile', None)
        if profile is None or not profile.is_manager:
            raise PermissionDenied('Hanya Manager Area yang dapat mengakses halaman ini.')
        return view_func(request, *args, **kwargs)
    return wrapper


# =============================================================================
# HALAMAN PUBLIK (tanpa login) - untuk PELANGGAN
# =============================================================================
@login_required
def home_submission(request):
    """Form input komplain. HANYA bisa diakses oleh user berperan Staff Input Komplain.
    User lain yang login otomatis dialihkan ke Dashboard.
    Mendukung prefill via QR/URL: ?cabang=<kode_outlet>&meja=<nomor_meja>
    """
    profile = getattr(request.user, 'staff_profile', None)
    if profile is None or not profile.is_input_staff:
        return redirect('complaints:dashboard')

    initial = {}
    branch_code = request.GET.get('cabang')
    initial_branch = None
    if branch_code:
        initial_branch = Branch.objects.filter(code__iexact=branch_code, is_active=True).first()
        if initial_branch:
            initial['branch'] = initial_branch.pk
    initial['cs_handled_time'] = timezone.localtime(timezone.now()).strftime('%Y-%m-%dT%H:%M')

    if request.method == 'POST':
        form = ComplaintSubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            complaint = form.save()
            messages.success(
                request,
                f'Komplain Anda berhasil dikirim. Kode komplain Anda: {complaint.code}. '
                'Simpan kode ini untuk mengecek status.'
            )
            return redirect('complaints:submission_success', code=complaint.code)
    else:
        form = ComplaintSubmissionForm(initial=initial)

    context = {
        'form': form,
        'cities': City.objects.filter(is_active=True).order_by('name'),
        'branches': Branch.objects.filter(is_active=True).select_related('city').order_by('name'),
        'initial_branch_id': initial_branch.pk if initial_branch else None,
        'initial_city_id': initial_branch.city_id if initial_branch else None,
        'detail_items': ComplaintDetailItem.objects.filter(is_active=True).order_by('main_type', 'name'),
    }
    return render(request, 'complaints/home.html', context)


def submission_success(request, code):
    complaint = get_object_or_404(Complaint, code=code)
    return render(request, 'complaints/submission_success.html', {'complaint': complaint})


def status_check(request):
    """Pelanggan mengecek status komplain dengan kode + no HP."""
    complaint = None
    if request.method == 'POST':
        form = StatusCheckForm(request.POST)
        if form.is_valid():
            complaint = Complaint.objects.filter(
                code__iexact=form.cleaned_data['code'].strip(),
                customer_phone=form.cleaned_data['customer_phone'].strip(),
            ).first()
            if not complaint:
                messages.error(request, 'Komplain tidak ditemukan. Periksa kembali kode dan nomor HP Anda.')
    else:
        form = StatusCheckForm()

    return render(request, 'complaints/status_check.html', {'form': form, 'complaint': complaint})


def satisfaction_rating(request, code):
    """Pelanggan memberi rating kepuasan setelah komplain berstatus 'selesai'."""
    complaint = get_object_or_404(Complaint, code=code)

    if complaint.status != Complaint.Status.SELESAI:
        messages.info(request, 'Rating hanya dapat diberikan setelah komplain selesai ditangani.')
        return redirect('complaints:status_check')

    if request.method == 'POST':
        form = SatisfactionRatingForm(request.POST, instance=complaint)
        if form.is_valid():
            rated = form.save(commit=False)
            rated.rated_at = timezone.now()
            rated.save()
            messages.success(request, 'Terima kasih atas penilaian Anda!')
            return redirect('complaints:status_check')
    else:
        form = SatisfactionRatingForm(instance=complaint)

    return render(request, 'complaints/rating.html', {'form': form, 'complaint': complaint})


# =============================================================================
# LOGIN / LOGOUT - untuk Staff/PIC, Manager, Admin Pusat
# =============================================================================
class StaffLoginView(auth_views.LoginView):
    template_name = 'complaints/login.html'


class StaffLogoutView(auth_views.LogoutView):
    pass


# =============================================================================
# GANTI PASSWORD - bisa dipakai SEMUA user yang sudah login (peran apa pun)
# =============================================================================
class StaffPasswordChangeView(auth_views.PasswordChangeView):
    template_name = 'complaints/password_change_form.html'
    form_class = StyledPasswordChangeForm
    success_url = reverse_lazy('complaints:password_change_done')


class StaffPasswordChangeDoneView(auth_views.PasswordChangeDoneView):
    template_name = 'complaints/password_change_done.html'


# =============================================================================
# Helper: batasi queryset komplain sesuai peran user
# =============================================================================
def _visible_complaints_for(user):
    qs = Complaint.objects.select_related('branch', 'assigned_to')
    profile = getattr(user, 'staff_profile', None)
    if profile is None:
        return qs.none()
    if profile.has_full_visibility:
        # Admin Pusat, Manager Wilayah, dan Staff Input Komplain melihat SEMUA outlet.
        return qs
    if profile.is_manager or profile.is_qc_trainer:
        # Manager Kota & QC/Trainer: semua outlet yang berada di kota yang sama.
        if profile.city_id:
            return qs.filter(branch__city=profile.city)
        return qs.none()
    if profile.is_staff_pic:
        # Staff/PIC Outlet: hanya outlet spesifiknya sendiri.
        if profile.branch_id:
            return qs.filter(branch=profile.branch)
        return qs.none()
    return qs.none()


# =============================================================================
# DASHBOARD (Staff / Manager / Admin Pusat)
# =============================================================================
@login_required
def dashboard(request):
    qs = _visible_complaints_for(request.user)

    # Filter Kota (khusus role yang punya akses ke semua kota, mis. Pusat/Admin/CS)
    profile = getattr(request.user, 'staff_profile', None)
    available_cities = City.objects.none()
    selected_city_id = request.GET.get('kota') or ''
    selected_city = None
    if profile and profile.has_full_visibility:
        available_cities = City.objects.filter(is_active=True).order_by('name')
        if selected_city_id:
            qs = qs.filter(branch__city_id=selected_city_id)
            selected_city = available_cities.filter(pk=selected_city_id).first()

    stats = {
        'total': qs.count(),
        'baru': qs.filter(status=Complaint.Status.BARU).count(),
        'diproses': qs.filter(
            status__in=[Complaint.Status.DITINJAU, Complaint.Status.DIPROSES]
        ).count(),
        'selesai': qs.filter(status=Complaint.Status.SELESAI).count(),
        'overdue': sum(1 for c in qs.exclude(
            status__in=[Complaint.Status.SELESAI, Complaint.Status.DITOLAK]
        ) if c.is_overdue),
        'avg_rating': qs.filter(satisfaction_rating__isnull=False).aggregate(
            avg=Avg('satisfaction_rating'))['avg'],
    }

    by_category = qs.values('category').annotate(total=Count('id')).order_by('-total')
    by_branch = qs.values('branch__name').annotate(total=Count('id')).order_by('-total')

    # ------------------------------------------------------------------
    # Ranking Outlet: jumlah komplain & jumlah lewat SLA per outlet
    # ------------------------------------------------------------------
    now = timezone.now()
    overdue_filter = Q(sla_deadline__lt=now) & ~Q(
        status__in=[Complaint.Status.SELESAI, Complaint.Status.DITOLAK]
    )
    outlet_ranking_qs = qs.values('branch__name').annotate(
        total=Count('id'),
        overdue=Count('id', filter=overdue_filter),
    ).order_by('-total')
    outlet_ranking = [
        {'name': row['branch__name'] or '-', 'total': row['total'], 'overdue': row['overdue']}
        for row in outlet_ranking_qs
    ]

    city_ranking_qs = qs.values('branch__city__name').annotate(
        total=Count('id'),
        overdue=Count('id', filter=overdue_filter),
    ).order_by('-total')
    city_ranking = [
        {'name': row['branch__city__name'] or '-', 'total': row['total'], 'overdue': row['overdue']}
        for row in city_ranking_qs
    ]

    recent_complaints = qs.order_by('-created_at')[:8]

    # ------------------------------------------------------------------
    # Data untuk grafik (Chart.js) di dashboard
    # ------------------------------------------------------------------
    # 1) Tren jumlah komplain per hari, 14 hari terakhir
    today = timezone.localdate()
    start_day = today - timedelta(days=13)
    trend_qs = (
        qs.filter(created_at__date__gte=start_day)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Count('id'))
    )
    trend_by_day = {row['day']: row['total'] for row in trend_qs}
    trend_labels = []
    trend_values = []
    for i in range(14):
        day = start_day + timedelta(days=i)
        trend_labels.append(day.strftime('%d %b'))
        trend_values.append(trend_by_day.get(day, 0))

    # 2) Distribusi status
    status_labels = [label for _, label in Complaint.Status.choices]
    status_counts_map = {row['status']: row['total'] for row in qs.values('status').annotate(total=Count('id'))}
    status_values = [status_counts_map.get(value, 0) for value, _ in Complaint.Status.choices]

    # 3) Distribusi tingkat keparahan
    severity_labels = [label for _, label in Complaint.Severity.choices]
    severity_counts_map = {row['severity']: row['total'] for row in qs.values('severity').annotate(total=Count('id'))}
    severity_values = [severity_counts_map.get(value, 0) for value, _ in Complaint.Severity.choices]

    # 4) Distribusi kategori (untuk grafik batang)
    category_labels = [dict(Complaint.Category.choices).get(row['category'], row['category']) for row in by_category]
    category_values = [row['total'] for row in by_category]

    chart_data = {
        'trend': {'labels': trend_labels, 'values': trend_values},
        'status': {'labels': status_labels, 'values': status_values},
        'severity': {'labels': severity_labels, 'values': severity_values},
        'category': {'labels': category_labels, 'values': category_values},
    }

    context = {
        'stats': stats,
        'by_category': by_category,
        'by_branch': by_branch,
        'outlet_ranking': outlet_ranking,
        'city_ranking': city_ranking,
        'recent_complaints': recent_complaints,
        'chart_data': chart_data,
        'available_cities': available_cities,
        'selected_city_id': selected_city_id,
        'selected_city': selected_city,
    }
    return render(request, 'complaints/dashboard.html', context)


# =============================================================================
# DAFTAR & DETAIL KOMPLAIN (internal)
# =============================================================================
@login_required
def complaint_list(request):
    qs = _visible_complaints_for(request.user)

    status = request.GET.get('status')
    severity = request.GET.get('severity')
    search = request.GET.get('q')
    only_overdue = request.GET.get('overdue')

    if status:
        qs = qs.filter(status=status)
    if severity:
        qs = qs.filter(severity=severity)
    if search:
        qs = qs.filter(
            Q(code__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(customer_phone__icontains=search) |
            Q(description__icontains=search)
        )

    qs = qs.order_by('-created_at')

    if only_overdue:
        qs = [c for c in qs if c.is_overdue]

    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'status_choices': Complaint.Status.choices,
        'severity_choices': Complaint.Severity.choices,
        'current_status': status or '',
        'current_severity': severity or '',
        'search': search or '',
        'only_overdue': bool(only_overdue),
    }
    return render(request, 'complaints/complaint_list.html', context)


@login_required
def complaint_detail(request, pk):
    complaint = get_object_or_404(_visible_complaints_for(request.user), pk=pk)
    profile = getattr(request.user, 'staff_profile', None)

    if request.method == 'POST':
        # Simpan kondisi SEBELUM form memodifikasi instance, untuk mendeteksi
        # transisi "baru pertama kali terisi" secara akurat.
        was_resolution_filled = bool(complaint.resolution_notes)
        was_internal_filled = bool(complaint.internal_notes)
        was_solution_confirmed = complaint.solution_confirmed
        was_validated = complaint.validated

        form = ComplaintUpdateForm(request.POST, request.FILES, instance=complaint, profile=profile)
        if form.is_valid():
            updated = form.save(commit=False)

            # "Ditangani oleh" otomatis terisi siapapun yang menyimpan perubahan.
            if profile is not None:
                updated.assigned_to = request.user

            # Transisi status otomatis: Staff/QC-Trainer mengisi Akar Masalah -> Ditinjau
            if profile and profile.can_handle_case:
                if updated.resolution_notes and not was_resolution_filled:
                    updated.status = Complaint.Status.DITINJAU
                # Staff/QC-Trainer mengisi Solusi -> Diproses
                if updated.internal_notes and not was_internal_filled:
                    updated.status = Complaint.Status.DIPROSES

            # [Gerbang 1] Validator mencentang "Solusi sudah benar" -> TIDAK
            # mengubah status, hanya membuka akses Staff untuk mengisi Validasi.
            if profile and profile.is_validator:
                if updated.solution_confirmed and not was_solution_confirmed:
                    updated.solution_confirmed_by = request.user
                    updated.solution_confirmed_at = timezone.now()

            # [Gerbang 2] Validator mencentang "Validasi sudah benar" -> Selesai
            if profile and profile.is_validator:
                if updated.validated and not was_validated:
                    updated.status = Complaint.Status.SELESAI
                    updated.resolved_at = timezone.now()
                    updated.validated_by = request.user
                    updated.validated_at = timezone.now()

            # Pengaman: status HARUS "Selesai" setiap kali validated=True tersimpan,
            # supaya tidak pernah "nyangkut" di status lain karena urutan aksi yang tidak biasa.
            if updated.validated:
                updated.status = Complaint.Status.SELESAI
                if not updated.resolved_at:
                    updated.resolved_at = timezone.now()

            # Tanggapan MA: catat siapa & kapan terakhir mengisi/mengubahnya.
            if profile and profile.is_manager and updated.manager_response:
                updated.manager_response_by = request.user
                updated.manager_response_at = timezone.now()

            # Tindak Lanjut LO: catat siapa & kapan terakhir mengisi/mengubahnya.
            if profile and profile.is_staff_pic and updated.lo_followup:
                updated.lo_followup_by = request.user
                updated.lo_followup_at = timezone.now()

            updated.save()
            messages.success(request, f'Komplain {complaint.code} berhasil diperbarui.')
            return redirect('complaints:complaint_detail', pk=pk)
    else:
        form = ComplaintUpdateForm(instance=complaint, profile=profile)

    show_penanganan_card = bool(form.fields) or (
        profile and complaint.internal_notes and (profile.is_validator or profile.can_handle_case)
    )

    context = {
        'complaint': complaint,
        'form': form,
        'timeline': complaint.timeline.all(),
        'can_edit': show_penanganan_card,
    }
    return render(request, 'complaints/complaint_detail.html', context)


# =============================================================================
# EXPORT KE EXCEL
# =============================================================================
@login_required
def export_complaints_excel(request):
    """Export seluruh data komplain (sesuai hak akses & filter aktif) ke file Excel."""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    qs = _visible_complaints_for(request.user)

    # Terapkan filter yang sama dengan halaman Daftar Komplain, jika ada di query string
    status = request.GET.get('status')
    severity = request.GET.get('severity')
    search = request.GET.get('q')
    only_overdue = request.GET.get('overdue')
    kota = request.GET.get('kota')

    if status:
        qs = qs.filter(status=status)
    if kota:
        qs = qs.filter(branch__city_id=kota)
    if severity:
        qs = qs.filter(severity=severity)
    if search:
        qs = qs.filter(
            Q(code__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(customer_phone__icontains=search) |
            Q(description__icontains=search)
        )

    qs = qs.order_by('-created_at')
    if only_overdue:
        qs = [c for c in qs if c.is_overdue]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Komplain'

    headers = [
        'Kode', 'Nama Pelanggan', 'No. HP', 'Email', 'Outlet', 'No. Meja',
        'Tanggal Kunjungan', 'No. Pesanan', 'Sumber Komplain',
        'Jam Komplain Masuk', 'Jam Ditangani CS',
        'Jenis Komplain', 'Rincian Komplain', 'Tingkat Keparahan',
        'Status', 'Deskripsi', 'Ditangani Oleh', 'Akar Masalah',
        'Dilaporkan Pada', 'Batas SLA', 'Lewat SLA?', 'Selesai Pada',
        'Rating Kepuasan', 'Masukan Tambahan',
    ]

    header_fill = PatternFill(start_color='7A1F1F', end_color='7A1F1F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, c in enumerate(qs, start=2):
        ws.cell(row=row_idx, column=1, value=c.code)
        ws.cell(row=row_idx, column=2, value=c.customer_name)
        ws.cell(row=row_idx, column=3, value=c.customer_phone)
        ws.cell(row=row_idx, column=4, value=c.customer_email)
        ws.cell(row=row_idx, column=5, value=c.branch.name if c.branch else '')
        ws.cell(row=row_idx, column=6, value=c.table_number)
        ws.cell(row=row_idx, column=7, value=c.visit_date.strftime('%d-%m-%Y') if c.visit_date else '')
        ws.cell(row=row_idx, column=8, value=c.order_number)
        ws.cell(row=row_idx, column=9, value=str(c.source) if c.source else '')
        ws.cell(row=row_idx, column=10,
                value=timezone.localtime(c.customer_complaint_time).strftime('%d-%m-%Y %H:%M') if c.customer_complaint_time else '')
        ws.cell(row=row_idx, column=11,
                value=timezone.localtime(c.cs_handled_time).strftime('%d-%m-%Y %H:%M') if c.cs_handled_time else '')
        ws.cell(row=row_idx, column=12, value=c.get_category_display())
        ws.cell(row=row_idx, column=13, value=c.detail_item.name if c.detail_item else '')
        ws.cell(row=row_idx, column=14, value=c.get_severity_display())
        ws.cell(row=row_idx, column=15, value=c.get_status_display())
        ws.cell(row=row_idx, column=16, value=c.description)
        ws.cell(row=row_idx, column=17, value=str(c.assigned_to) if c.assigned_to else '')
        ws.cell(row=row_idx, column=18, value=c.resolution_notes)
        ws.cell(row=row_idx, column=19,
                value=timezone.localtime(c.created_at).strftime('%d-%m-%Y %H:%M') if c.created_at else '')
        ws.cell(row=row_idx, column=20,
                value=timezone.localtime(c.sla_deadline).strftime('%d-%m-%Y %H:%M') if c.sla_deadline else '')
        ws.cell(row=row_idx, column=21, value='Ya' if c.is_overdue else 'Tidak')
        ws.cell(row=row_idx, column=22,
                value=timezone.localtime(c.resolved_at).strftime('%d-%m-%Y %H:%M') if c.resolved_at else '')
        ws.cell(row=row_idx, column=23, value=c.satisfaction_rating)
        ws.cell(row=row_idx, column=24, value=c.satisfaction_feedback)

    # Lebar kolom otomatis (sederhana, dibatasi agar tidak terlalu lebar)
    widths = [12, 20, 15, 22, 22, 10, 16, 16, 16, 18, 18, 16, 20, 16, 14, 40, 18, 35, 18, 18, 12, 18, 14, 35]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'data_komplain_{timezone.localdate().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# PANEL ADMIN PUSAT: kelola akun staff, outlet, dan identitas perusahaan
# =============================================================================
@admin_pusat_required
def admin_panel(request):
    context = {
        'total_users': User.objects.count(),
        'total_branches': Branch.objects.count(),
        'total_cities': City.objects.count(),
    }
    return render(request, 'complaints/admin_panel.html', context)


@admin_pusat_required
def user_list(request):
    users = User.objects.select_related(
        'staff_profile', 'staff_profile__branch', 'staff_profile__city'
    ).order_by('username')
    return render(request, 'complaints/user_list.html', {'users': users})


@admin_pusat_required
def user_create(request):
    if request.method == 'POST':
        form = StaffAccountCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Akun "{user.username}" berhasil dibuat.')
            return redirect('complaints:user_list')
        else:
            messages.error(request, 'Akun GAGAL dibuat. Periksa kesalahan yang ditandai di bawah ini.')
    else:
        form = StaffAccountCreateForm()
    return render(request, 'complaints/user_form.html', {'form': form, 'is_create': True})


@admin_pusat_required
def user_edit(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = StaffAccountEditForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Akun "{user_obj.username}" berhasil diperbarui.')
            return redirect('complaints:user_list')
        else:
            messages.error(request, 'Perubahan GAGAL disimpan. Periksa kesalahan yang ditandai di bawah ini.')
    else:
        form = StaffAccountEditForm(instance=user_obj)
    return render(request, 'complaints/user_form.html', {
        'form': form, 'is_create': False, 'user_obj': user_obj,
    })


@admin_pusat_required
def user_reset_password(request, pk):
    """Admin Pusat me-reset password user manapun ke password acak baru.
    Password ASLI tidak pernah bisa dilihat (disimpan dalam bentuk hash satu-arah,
    ini standar keamanan universal) -- ini alternatif yang aman: generate password
    baru dan tampilkan SATU KALI di layar supaya bisa disalin & diberikan ke staf.
    """
    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        new_password = secrets.token_urlsafe(9)  # ~12 karakter, mudah dibaca & disalin
        user_obj.set_password(new_password)
        user_obj.save(update_fields=['password'])
        messages.success(
            request,
            f'Password baru untuk "{user_obj.username}" adalah: {new_password} '
            '(salin sekarang, tidak akan ditampilkan lagi setelah ini).'
        )
        return redirect('complaints:user_list')
    return render(request, 'complaints/user_reset_password_confirm.html', {'user_obj': user_obj})


@admin_pusat_required
def branch_list(request):
    branches = Branch.objects.select_related('manager').order_by('name')
    return render(request, 'complaints/branch_list.html', {'branches': branches})


@admin_pusat_required
def branch_create(request):
    if request.method == 'POST':
        form = BranchAdminForm(request.POST)
        if form.is_valid():
            branch = form.save()
            messages.success(request, f'Outlet "{branch.name}" berhasil ditambahkan.')
            return redirect('complaints:branch_list')
    else:
        form = BranchAdminForm()
    return render(request, 'complaints/branch_form.html', {'form': form, 'is_create': True})


@admin_pusat_required
def branch_edit(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == 'POST':
        form = BranchAdminForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, f'Outlet "{branch.name}" berhasil diperbarui.')
            return redirect('complaints:branch_list')
    else:
        form = BranchAdminForm(instance=branch)
    return render(request, 'complaints/branch_form.html', {
        'form': form, 'is_create': False, 'branch': branch,
    })


@admin_pusat_required
def site_settings_edit(request):
    settings_obj = SiteSettings.load()
    if request.method == 'POST':
        form = SiteSettingsForm(request.POST, request.FILES, instance=settings_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Identitas perusahaan berhasil diperbarui.')
            return redirect('complaints:site_settings_edit')
    else:
        form = SiteSettingsForm(instance=settings_obj)
    return render(request, 'complaints/site_settings_form.html', {
        'form': form, 'settings_obj': settings_obj,
    })


# =============================================================================
# PANEL ADMIN PUSAT: kelola KOTA (hanya Admin Pusat yang boleh menambah/mengubah)
# =============================================================================
@admin_pusat_required
def city_list(request):
    cities = City.objects.all().order_by('name')
    return render(request, 'complaints/city_list.html', {'cities': cities})


@admin_pusat_required
def city_create(request):
    if request.method == 'POST':
        form = CityAdminForm(request.POST)
        if form.is_valid():
            city = form.save()
            messages.success(request, f'Kota "{city.name}" berhasil ditambahkan.')
            return redirect('complaints:city_list')
    else:
        form = CityAdminForm()
    return render(request, 'complaints/city_form.html', {'form': form, 'is_create': True})


@admin_pusat_required
def city_edit(request, pk):
    city = get_object_or_404(City, pk=pk)
    if request.method == 'POST':
        form = CityAdminForm(request.POST, instance=city)
        if form.is_valid():
            form.save()
            messages.success(request, f'Kota "{city.name}" berhasil diperbarui.')
            return redirect('complaints:city_list')
    else:
        form = CityAdminForm(instance=city)
    return render(request, 'complaints/city_form.html', {
        'form': form, 'is_create': False, 'city': city,
    })


# =============================================================================
# PANEL ADMIN PUSAT: kelola SUMBER KOMPLAIN (hanya Admin Pusat)
# =============================================================================
@admin_pusat_required
def source_list(request):
    sources = ComplaintSource.objects.all().order_by('name')
    return render(request, 'complaints/source_list.html', {'sources': sources})


@admin_pusat_required
def source_create(request):
    if request.method == 'POST':
        form = ComplaintSourceForm(request.POST)
        if form.is_valid():
            source = form.save()
            messages.success(request, f'Sumber Komplain "{source.name}" berhasil ditambahkan.')
            return redirect('complaints:source_list')
    else:
        form = ComplaintSourceForm()
    return render(request, 'complaints/source_form.html', {'form': form, 'is_create': True})


@admin_pusat_required
def source_edit(request, pk):
    source = get_object_or_404(ComplaintSource, pk=pk)
    if request.method == 'POST':
        form = ComplaintSourceForm(request.POST, instance=source)
        if form.is_valid():
            form.save()
            messages.success(request, f'Sumber Komplain "{source.name}" berhasil diperbarui.')
            return redirect('complaints:source_list')
    else:
        form = ComplaintSourceForm(instance=source)
    return render(request, 'complaints/source_form.html', {
        'form': form, 'is_create': False, 'source': source,
    })


# =============================================================================
# PANEL ADMIN PUSAT: kelola RINCIAN KOMPLAIN (hanya Admin Pusat)
# =============================================================================
@admin_pusat_required
def detail_item_list(request):
    items = ComplaintDetailItem.objects.all().order_by('main_type', 'name')
    return render(request, 'complaints/detail_item_list.html', {'items': items})


@admin_pusat_required
def detail_item_create(request):
    if request.method == 'POST':
        form = ComplaintDetailItemForm(request.POST)
        if form.is_valid():
            item = form.save()
            messages.success(request, f'Rincian Komplain "{item.name}" berhasil ditambahkan.')
            return redirect('complaints:detail_item_list')
    else:
        form = ComplaintDetailItemForm()
    return render(request, 'complaints/detail_item_form.html', {'form': form, 'is_create': True})


@admin_pusat_required
def detail_item_edit(request, pk):
    item = get_object_or_404(ComplaintDetailItem, pk=pk)
    if request.method == 'POST':
        form = ComplaintDetailItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'Rincian Komplain "{item.name}" berhasil diperbarui.')
            return redirect('complaints:detail_item_list')
    else:
        form = ComplaintDetailItemForm(instance=item)
    return render(request, 'complaints/detail_item_form.html', {
        'form': form, 'is_create': False, 'item': item,
    })


# =============================================================================
# MANAGER AREA: kelola nomor WhatsApp QC/Trainer & Leader Outlet di kotanya
# =============================================================================
@manager_required
def manager_staff_list(request):
    profile = request.user.staff_profile
    city = profile.city

    if city:
        staff_profiles = StaffProfile.objects.filter(
            Q(role=StaffProfile.Role.QC_TRAINER, city=city) |
            Q(role=StaffProfile.Role.STAFF, branch__city=city)
        ).select_related('user', 'branch', 'city').order_by('role', 'user__username')
    else:
        staff_profiles = StaffProfile.objects.none()

    return render(request, 'complaints/manager_staff_list.html', {
        'staff_profiles': staff_profiles, 'city': city,
    })


@manager_required
def manager_edit_phone(request, pk):
    profile = request.user.staff_profile
    city = profile.city
    target = get_object_or_404(StaffProfile, pk=pk)

    # Pastikan target BENAR-BENAR dalam cakupan kota Manager Area ini,
    # supaya tidak bisa mengubah nomor QC/Trainer atau Leader Outlet kota lain.
    in_scope = False
    if city:
        if target.role == StaffProfile.Role.QC_TRAINER and target.city_id == city.id:
            in_scope = True
        elif target.role == StaffProfile.Role.STAFF and target.branch and target.branch.city_id == city.id:
            in_scope = True

    if not in_scope:
        raise PermissionDenied('Anda tidak berhak mengubah data user ini.')

    if request.method == 'POST':
        form = StaffPhoneEditForm(request.POST, instance=target)
        if form.is_valid():
            form.save()
            messages.success(request, f'Nomor WhatsApp "{target.user.username}" berhasil diperbarui.')
            return redirect('complaints:manager_staff_list')
    else:
        form = StaffPhoneEditForm(instance=target)

    return render(request, 'complaints/manager_edit_phone.html', {
        'form': form, 'target': target,
    })
